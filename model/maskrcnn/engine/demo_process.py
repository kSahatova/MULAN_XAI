# Ke Yan, Imaging Biomarkers and Computer-Aided Diagnosis Laboratory,
# National Institutes of Health Clinical Center, July 2019
"""Procedure in the demo mode"""
import os
import pandas as pd
import numpy as np
from time import time
import torch
import nibabel as nib
from tqdm import tqdm
import cv2
from openpyxl import load_workbook

from maskrcnn.config import cfg
from maskrcnn.data.datasets.load_ct_img import load_prep_img
from maskrcnn.structures.image_list import to_image_list
from maskrcnn.data.datasets.evaluation.DeepLesion.post_process import post_process_results
from maskrcnn.data.datasets.load_ct_img import windowing, windowing_rev
from maskrcnn.utils.draw import draw_results


def exec_model(model, path, window_indx=None):
    """test model on user-provided data, instead of the preset DeepLesion dataset"""
    import_tag_data()
    model.eval()
    device = torch.device(cfg.MODEL.DEVICE)

    nifti_data = None
    if not os.path.exists(path):
        print('file does not exist!')
        return
    try:
        print('reading image ...')
        nifti_data = nib.load(path)
    except:
        print('load nifti file error!')
        return
    
    if window_indx is None:
        window_indx = input('Window to show, 1:soft tissue, 2:lung, 3: bone >> ')
    if window_indx not in ['1', '2', '3']:
        print('Please, provide correct window index')
    win_sel = window_indx
    win_show = [[-175, 275], [-1500, 500], [-500, 1300]]
    win_show = win_show[int(win_sel)-1]

    vol, spacing, slice_intv = load_preprocess_nifti(nifti_data)
    slice_num_per_run = max(1, int(float(cfg.TEST.TEST_SLICE_INTV_MM)/slice_intv+.5))
    num_total_slice = vol.shape[2]

    total_time = 0
    output_dir = os.path.join(cfg.RESULTS_DIR,path.replace(os.sep, '_'))
    if not os.path.exists(output_dir):
        os.mkdir(output_dir)

    slices_to_process = range(int(slice_num_per_run/2), num_total_slice, slice_num_per_run) #range(0, 3)
    msgs_all = []
    print('predicting ...')
    slices_result = []
    img_slices = []
    print('slices_to_process', slices_to_process)
    for slice_idx in tqdm(slices_to_process):  
        ims, im_np, im_scale, crop = get_ims(slice_idx, vol, spacing, slice_intv)
        im_list = to_image_list(ims, cfg.DATALOADER.SIZE_DIVISIBILITY).to(device)
        #print(im_list[0].tensors.shape)
        img_slices.append(im_list)
        
        start_time = time()
        with torch.no_grad():
            result = model(im_list)
        
        result = [o.to("cpu") for o in result]
        
        info = {'spacing': spacing, 'im_scale': im_scale}
        post_process_results(result[0], info)
        total_time += time() - start_time
        
        output_fn = os.path.join(output_dir, '%d.png'%(slice_idx+1))
        overlay, msgs, results = gen_output(im_np, result[0], info, win_show)

        slices_result.append(results)

        cv2.imwrite(output_fn, overlay)
        msgs_all.append('slice %d\r\n' % (slice_idx+1))
        for msg in msgs:
            msgs_all.append(msg+'\r\n')
        msgs_all.append('\r\n')

    with open(os.path.join(output_dir, 'results.txt'), 'w') as f:
        f.writelines(msgs_all)

    print('result images and text saved to', output_dir)
    print('processing time: %d ms per slice' % int(1000.*total_time/len(slices_to_process)))
        
    return slices_result, img_slices


def get_windows_inds(anno_file_path, img_paths):
    '''Extracts DICOM window size according to the annotation file
        Args:
            anno_file_path: str - path to the annotation file
            img_paths: list[str] or str - path to the input image
    '''
    
    anno_file = pd.read_csv(anno_file_path)
    win_inds = []
    
    if not isinstance(img_paths, (list, tuple)):
        img_paths = list(img_paths)
        
    for img_path in img_paths:
        case_ind = img_path.rsplit('_', 1)[0]
        win_sizes = anno_file[anno_file.File_name.str.startswith(case_ind)]['DICOM_windows']
        win_sizes = win_sizes.tolist()[0]    
        if win_sizes == '-175, 275':
            win_ind = '1'
        elif win_sizes == '-1500, 500':
            win_ind = '2'
        elif win_sizes == '-500, 1300':
            win_ind = '3'
        win_inds.append(win_ind)
        
    return win_inds
     
    
def load_nifti_file(path, device=torch.device('cuda:0')):
    '''Returns list of ImageLists loaded from nifti file'''
    
    nifti_data = nib.load(path)
    vol, spacing, slice_intv = load_preprocess_nifti(nifti_data)
    
    slice_num_per_run = max(1, int(float(cfg.TEST.TEST_SLICE_INTV_MM)/slice_intv+.5))
    num_total_slice = vol.shape[2]
    
    slices_to_process = range(int(slice_num_per_run/2), num_total_slice, slice_num_per_run)
    img_slices = []
    for slice_idx in tqdm(slices_to_process): 
        ims, im_np, im_scale, crop = get_ims(slice_idx, vol, spacing, slice_intv)
        im_list = to_image_list(ims, cfg.DATALOADER.SIZE_DIVISIBILITY).to(device)
        img_slices.append(im_list)        
        
    return img_slices


def import_tag_data():
    cellname = lambda row, col: '%s%d' % (chr(ord('A') + col - 1), row)
    fn = os.path.join(cfg.PROGDAT_DIR, '%s_%s.xlsx' % ('test_handlabeled', cfg.EXP_NAME))

    wb = load_workbook(fn)
    sheetnames = wb.sheetnames
    sheet = wb[sheetnames[0]]
    tags = []
    thresolds = []
    for p in range(2, sheet.max_row):
        tags.append(sheet[cellname(p, 1)].value)
        thresolds.append(float(sheet[cellname(p, 8)].value))
    assert tags == cfg.runtime_info.tag_list
    cfg.runtime_info.tag_sel_val = torch.tensor(thresolds).to(torch.float)


def load_preprocess_nifti(data):
    vol = (data.get_data().astype('int32') + 32768).astype('uint16')  # to be consistent with png files
    # spacing = -data.get_affine()[0,1]
    # slice_intv = -data.get_affine()[2,2]
    aff = data.affine[:3, :3]
    spacing = np.abs(aff[:2, :2]).max()
    slice_intv = np.abs(aff[2, 2])

    # TODO: Ad-hoc code for normalizing the orientation of the volume.
    # The aim is to make vol[:,:,i] an supine right-left slice
    # It works for the authors' data, but maybe not suitable for some kinds of nifti files
    if np.abs(aff[0, 0]) > np.abs(aff[0, 1]):
        vol = np.transpose(vol, (1, 0, 2))
        aff = aff[[1, 0, 2], :]
    if np.max(aff[0, :2]) > 0:
        vol = vol[::-1, :, :]
    if np.max(aff[1, :2]) > 0:
        vol = vol[:, ::-1, :]
    return vol, spacing, slice_intv


def get_ims(slice_idx, vol, spacing, slice_intv):
    num_slice = cfg.INPUT.NUM_SLICES * cfg.INPUT.NUM_IMAGES_3DCE
    im_np, im_scale, crop = load_prep_img(vol, slice_idx, spacing, slice_intv,
                                          cfg.INPUT.IMG_DO_CLIP, num_slice=num_slice)
    im = im_np - cfg.INPUT.PIXEL_MEAN
    im = torch.from_numpy(im.transpose((2, 0, 1))).to(dtype=torch.float)
    ims = im.split(cfg.INPUT.NUM_IMAGES_3DCE)
    return ims, im_np[:, :, int(num_slice/2)+1], im_scale, crop


def gen_output(im, result, info, win_show):
    im = windowing_rev(im, cfg.INPUT.WINDOWING)
    im = windowing(im, win_show).astype('uint8')
    im = cv2.cvtColor(im, cv2.COLOR_GRAY2RGB)

    scale = cfg.TEST.VISUALIZE.SHOW_SCALE
    im = cv2.resize(im, None, None, fx=scale, fy=scale, interpolation=cv2.INTER_LINEAR)

    pred = result.bbox.cpu().numpy()
    labels = result.get_field('labels').cpu().numpy()
    scores = result.get_field('scores').cpu().numpy()
    tag_scores = result.get_field('tag_scores').cpu().numpy()
    tag_predictions = result.get_field('tag_predictions').cpu().numpy()

    mm2pix = info['im_scale'] / info['spacing'] * scale
    contours = result.get_field('contour_mm').cpu().numpy() * mm2pix
    contours = [c[c[:, 0] > 0, :] for c in contours]
    contours = [c+1*scale for c in contours]  # there seems to be a small offset in the mask?
    recists = result.get_field('recist_mm').cpu().numpy() * mm2pix
    recists += 1*scale   # there seems to be a small offset in the mask?
    diameters = result.get_field('diameter_mm').cpu().numpy()
    results = {"bboxes": pred, "scores": scores,
               "labels": labels, "tag_scores": tag_scores, 
               "tag_predictions": tag_predictions, "recists": recists,
               "diameters": diameters, 'contours': contours}

    pred *= scale
    overlay, msgs = draw_results(im, pred, labels, scores, tag_predictions=tag_predictions, tag_scores=tag_scores,
                                 contours=contours, recists=recists, diameters=diameters)
    overlay = print_msg_on_img(overlay, msgs)
    return overlay, msgs, results


def print_msg_on_img(overlay, msgs):
    txt_height = 20
    msg_im = np.zeros((txt_height*cfg.TEST.VISUALIZE.DETECTIONS_PER_IMG+10, overlay.shape[1], 3), dtype=np.uint8)
    for p in range(len(msgs)):
        msg = msgs[p].split(' | ')
        msg = msg[0][7:10] + msg[1][:-2] + ': ' + msg[2]
        cv2.putText(msg_im, msg, (0, txt_height*(p+1)),
                    cv2.FONT_HERSHEY_DUPLEX, fontScale=.5,
                    color=(255,255,255), thickness=1)
    return np.vstack((overlay, msg_im))
